from datetime import date, datetime
import os
from openpyxl.styles import PatternFill, Border, Side, Alignment, Font
from openpyxl.drawing.image import Image
from openpyxl import load_workbook
from openpyxl import Workbook
import pandas as pd
import tempfile
import shutil
from utils import type_check


class ReportDesign:
    """
    This class is used to design either the SLA/Bot Report or the Home Delivery Report.

      Attributes:
        - sla_data (dict): SLA Data Dictionary
        - bot_df (Dataframe): Bot Data Frame
        - day_sorter (int): Day value that was used to filter the "Days" column.
        - highest_day (int): Highest value in the "Day" Column
        - shipped_awb_df (Dataframe): Shipped AWB Dataframe
        - non_shipped_df (Dataframe): Non-Shipped AWB Dataframe
      Methods:
        - create_report: Creates SLA/Bot or Home Delivery Report
        - set_column_widths: Set column width
        - create_full_borders: Sets borders around a cell.
        - change_font: Change font of a cell.
        - fill_color: Change fill color of a cell.
        - hide_gridlines: Hide gridlines.
        - all_cell_styles: Style all cells.
        - get_cell_coordinate: Get cell coordinates based on text.
        - get_date_time: Get datetime now
        - create_folder: Create a folder in current directory.
        - invalid_name: Checks if a file, folder or sheet name contains invalid characters.
    """

    # Valid Report constant which contains the report name and a tuple of the method name to be called along with the
    # instance attributes to be created at runtime.
    VALID_REPORT_DESIGN = {
        "SLA/Bot Report": ("_create_bot_sla_report", ("sla_data", "bot_df", "day_sorter", "highest_day")),
        "Home Delivery Report": ("_create_home_delivery_report", ("shipped_awb_df", "non_shipped_awb_df"))
    }

    def __init__(self, report_name: str):
        """
        Initializes a ReportName Object.

        Creates a new ReportDesign object with the specified report name.
        The `report_name` argument is used to determine which
        report to generate, and the appropriate method name and instance variables are
        retrieved from the `VALID_REPORT_DESIGN` dictionary. The instance variables are then
        created and initialized to `None` using the `setattr` method.

        :param report_name: The name of the report to be created. Must be one of the valid report names defined
            in ReportDesign.VALID_REPORT_DESIGN

        :raise KeyError: If the specified report name is not one of the valid report names defined in
            ReportDesign.VALID_REPORT_DESIGN.
        """
        self.temp_file = None
        self.workbook = None
        self.sheet = None

        if report_name not in ReportDesign.VALID_REPORT_DESIGN.keys():
            raise KeyError(f"{report_name} is not a valid report. Valid reports are "
                           f"{' or '.join(ReportDesign.VALID_REPORT_DESIGN.keys())}")

        # create_report stores the tuple values for the method name to call.
        # instance_variables stores the tuple values to create the necessary instance variables.
        report_name, instance_variables = ReportDesign.VALID_REPORT_DESIGN[report_name]

        # Loop through the instance_variables assign them self, the variable name and the value none.
        for var in instance_variables:
            setattr(self, var, None)

    def create_report(self, report_name: str) -> None:
        """
        Creates a report based on the specific name.

        Executes the appropriate method call based on the report name. Either _create_bot_sla_report or
        _create_home_delivery_report will be executed.

        :param report_name:  The name of the report to create.
        :raise KeyError: If the specified report name is not one of the valid report names defined in
            ReportDesign.VALID_REPORT_DESIGN.
        """
        if report_name not in ReportDesign.VALID_REPORT_DESIGN.keys():
            raise KeyError(f"{report_name} is not a valid report. Pass in the same Report Name as you did when you"
                           f" created the TableData Class. Valid reports are "
                           f"{' or '.join(ReportDesign.VALID_REPORT_DESIGN.keys())}")

        name_of_report = ReportDesign.VALID_REPORT_DESIGN[report_name][0]

        # Store the method object in report_method and then call that method based on the report_name passed in.
        report_method = getattr(self, name_of_report)
        report_method()

    def set_column_widths(self, column_widths: dict) -> None:
        """
        Set column width from a dictionary of columns.

        :param column_widths: Pass in a dictionary where the Keys are the columns (Ex. A, B) and the values are
            the width you want to set that column to (Ex. 50). (Ex. column_custom_width = {"A": 100, "B": 50}

        :raise ValueError: Will raise an error if the dictionary key is not a letter or if the length of the key is
            greater than 1. Also, will raise a KeyError if the width value in the dictionary is not an int.
        """
        for column_name, width in column_widths.items():
            type_check(arg=width, arg_name="width", expected_type=int)
            if not column_name.isalpha() or len(column_name) != 1:
                raise ValueError(f"Please ensure when passing column_custom_width that the Keys are"
                                 f" the name of the column. (Ex. 'A' or 'B' or 'C'")
            else:
                self.sheet.column_dimensions[column_name].width = width

    def create_full_borders(self, cell_coordinate: str, border_type: str = "thin") -> None:
        """
        Create a full border around a cell.
        :param cell_coordinate: Coordinate of the cell.
        :param border_type: Set the type of border. (Default: "thin")
        """
        self.sheet[cell_coordinate].border = Border(left=Side(border_type), right=Side(border_type),
                                                    top=Side(border_type), bottom=Side(border_type))

    def change_font(self, cell_coordinate: str, font_name: str = "Calibri", size: int = 11, bold: bool = False,
                    italics: bool = False, hex_color: str = "ffffff") -> None:
        """
        Change the font of a cell.
        :param cell_coordinate: Coordinate of the cell you want to change the font for. (Ex. E10)
        :param font_name: Name of the font (Default: 'Calibri')
        :param size: Size of the font (Default: 11)
        :param bold: Set bold font. True = Bold | False = Not Bold (Default: False)
        :param italics: Set italics font. True = Italic | False = Not Italic (Default: False)
        :param hex_color: Color you want the font to be. Make sure its in hex value without '#'.
            (Default = 'ffffff' [White Color])
        """
        self.sheet[cell_coordinate].font = Font(name=font_name, size=size, bold=bold, color=hex_color, italic=italics)

    def fill_color(self, cell_coordinate: str, hex_color: str, fill_type: str = "solid") -> None:
        """
        Change the fill color of a specific cell.
        :param cell_coordinate: Coordinate of the cell you want to fill. (Ex. E10)
        :param hex_color: Color you want to fill with. Make sure its in hex value without '#'. (Ex. 'ffffff')
        :param fill_type: Set a pattern or color gradient (Default: 'solid")
        """
        self.sheet[cell_coordinate].fill = PatternFill(start_color=hex_color, end_color=hex_color, fill_type=fill_type)

    def hide_gridlines(self, hide_gridlines: bool = True) -> None:
        """
        Hide/Show gridlines.
        :param hide_gridlines: Show/Hide the gridlines. (Default: 'True')
        :raise ValueError: Will raise error if gridlines is not a boolean value.
        """
        if hide_gridlines:
            self.sheet.sheet_view.showGridLines = False
        elif not hide_gridlines:
            self.sheet.sheet_view.showGridLines = True
        else:
            raise ValueError(f"{hide_gridlines} is not a valid option. Please only use 'True' or 'False'.")

    def all_cell_styles(self, horizontal_alignment: str = "center") -> None:
        """
        Changes all cells in a workbook.

        :param horizontal_alignment: Set horizontal alignment for all cells. (Default: "center")
        """
        for row in self.sheet.iter_rows():
            for cell in row:
                cell.alignment = Alignment(horizontal=horizontal_alignment)

    def get_cell_coordinate(self, cell_text: list) -> dict:
        """
        Get the cell coordinates based on text in a cell.

        The method will loop through the Excel document and find the text that matches the text in the list. It will
        store the key as the cell coordinate and the value will be the text inside that cell.
        :param cell_text: List of words to find where they are located in the Excel Document.
        :return: Returns the coordinates and values of the found text.
        """
        coordinate_dict = {cell.coordinate: cell.value for row in self.sheet.iter_rows() for cell in row
                           if cell.value in cell_text}

        return coordinate_dict

    def _insert_data_to_excel(self, dataframe: pd.DataFrame, start_col: int = 1, start_row: int = 1,
                             index: bool = True, header: bool = True, sheet_name: str = "Sheet") -> None:
        """
        Insert a Dataframe into a temporary Excel File.
        :param dataframe: Dataframe to insert.
        :param start_col: Column position to start inserting Dataframe (Default: 1).
        :param start_row: Row position to start inserting Dataframe (Default: 1)
        :param index: Display row names (Default: True)
        :param header: Display column names (Default: True)
        :param sheet_name: Name of the Sheet to insert Dataframe. (Default: "Sheet")
        :raise KeyError: Will raise error if sheet contains any invalid characters.
        """
        self.invalid_name(sheet_name=sheet_name)

        with pd.ExcelWriter(self.temp_file, mode="a", engine="openpyxl", if_sheet_exists="overlay") as writer:
            dataframe.to_excel(writer, startrow=start_row - 1, startcol=start_col - 1, header=header, index=index,
                               sheet_name=sheet_name)

    @staticmethod
    def get_date_time() -> str:
        """
        Get the current date and time.
        :return: Returns current date/time formatted in a string.
        """
        today_date = datetime.today()
        current_date_time = today_date.strftime("%B %d, %Y - %#I%M%p")
        return current_date_time

    @staticmethod
    def _get_total_weight(data_dictionary: dict) -> int:
        """
        Get the total of all values in a dictionary.
        :param data_dictionary: Dictionary in which you want to get the total values from.
        :return: Returns the sum of all values in a dictionary.
        """
        return sum(data_dictionary.values())

    @classmethod
    def _get_sla_bot_custom_widths(cls) -> dict:
        """
        Get specific column widths.
        :return: Returns a dictionary of column names as the keys and custom width as the values.
        """
        custom_column_width = {
            "B": 20,
            "C": 15,
            "E": 13,
            "F": 17,
            "G": 17,
            "H": 17,
            "I": 17,
            "J": 17,
            "K": 17,
            "L": 17,
            "M": 17,
        }

        return custom_column_width

    @classmethod
    def _get_non_shipped_awb_custom_widths(cls) -> dict:
        """
        Get specific column widths.
        :return: Returns a dictionary of column names as the keys and custom width as the values.
        """
        custom_column_width = {
            "B": 25,
            "C": 25,
            "D": 50,
        }

        return custom_column_width

    @classmethod
    def _get_shipped_awb_custom_widths(cls) -> dict:
        """
        Get specific column widths.
        :return: Returns a dictionary of column names as the keys and custom width as the values.
        """
        custom_column_width = {
            "B": 20,
            "C": 20,
            "D": 20,
            "E": 20,
            "F": 20,
            "G": 20,
            "H": 40,
        }

        return custom_column_width

    def _add_color_reference_header(self) -> None:
        """
        Add reference header cells.

        Add 4 Cells at the top of the Bot Report Table. The cells are just responsible for displaying
        color references. (Ex. F8 will contain "Going Today" with a fill color of green. Any cell in the Bot Report
        table which is colored green, means that cargo will be going today).
        """

        reference_data = {"E3": "Reference", "F3": "Going Today", "G3": "To Be Cleared", "H3": "On Hold"}
        for count, (cell_cord, cell_text) in enumerate(reference_data.items()):
            self.sheet[cell_cord].value = cell_text
            self.create_full_borders(cell_coordinate=cell_cord)
            if count == 0:
                self.change_font(cell_coordinate=cell_cord, bold=True)
            else:
                self.change_font(cell_coordinate=cell_cord, bold=True, hex_color="000000")

            # reference_data contains the cell coordinates as the keys. Convert it to a list to access those keys.
            reference_data_keys = list(reference_data.keys())
            self.fill_color(cell_coordinate=reference_data_keys[0], hex_color="4285f4")
            self.fill_color(cell_coordinate=reference_data_keys[1], hex_color="00ff00")
            self.fill_color(cell_coordinate=reference_data_keys[2], hex_color="ffff00")
            self.fill_color(cell_coordinate=reference_data_keys[3], hex_color="ff0000")

    def _add_days_top_pri_header(self) -> None:
        """
        Create the top header for Days and Top Priority.

        Add the Days Text and Priority Text. As well assign the value of each below there text.
        The day sorter is the value in the SLA/Bot Report setting which is used to filter data by the number of days
        cargo has been in the warehouse. TOP Pri, is the highest value in the "Days" column. All data
        is obtained from the table_data class. Method also formats this small table to the appropriate colors, borders
        and font.
        """
        day_pri_values = {"E5": "DAYS", "E6": self.day_sorter, "H5": "Days TOP PRI", "H6": self.highest_day}

        for cell_cord, cell_text in day_pri_values.items():
            self.sheet[cell_cord].value = cell_text
            self.change_font(cell_coordinate=cell_cord, bold=True)
            self.fill_color(cell_coordinate=cell_cord, hex_color="7A7A7A")

        self.fill_color(cell_coordinate="F5", hex_color="7A7A7A")
        self.sheet.merge_cells("F5:G6")

    def _add_date_header(self) -> None:
        """
        Add the current date.
        """
        today = date.today()
        formatted_date = today.strftime("%B %d, %Y")
        self.sheet["L5"].value = formatted_date
        self.change_font(cell_coordinate="L5", size=18, hex_color="000000")
        self.sheet["L5"].alignment = Alignment(vertical="center")
        self.sheet.merge_cells("L5:M6")

    def _add_logo(self, cell_coordinate) -> None:
        """
        Add logo.
        :param cell_coordinate: Cell where you want to place the logo.
        """
        img = Image("logo.png")
        self.sheet.add_image(img, cell_coordinate)

    def _create_sla_headers(self) -> None:
        """
        Create the SLA Table Header and designs it.
        """
        sla_header = {"Destination": "B8", "Past SLA": "C8"}
        self.sheet["B8"].value = "Destination"
        self.sheet["C8"].value = "Past SLA"

        self._common_header_design(data_dict=sla_header)

    def _add_sla_total_weight_header(self) -> int:
        """
        Create the Total | Weight header at the bottom of the SLA Table.
        :return: Returns the Total Weight row number as an int.
        """
        total_weight = ReportDesign._get_total_weight(self.sla_data)

        # Get the length of sla_data to tell us how many rows there are. Add 10 to that, to allow an empty row between
        # the data and the total weight. 10 is there since we start at row 9, then add 1 empty row so 9 + 1 = 10.
        total_sla_rows = len(self.sla_data) + 10
        self.sheet[f"B{total_sla_rows}"].value = "Total"
        self.sheet[f"C{total_sla_rows}"].value = total_weight

        return total_sla_rows

    def _sla_table_color(self, total_weight_row_num) -> None:
        """
        Adds the proper colors to the SLA Table

        :param total_weight_row_num: The row number for the Total Weight row.
        """
        # Start at row 9 and every second row, it will color those cells light blue.
        for row in range(9, len(self.sla_data) + 9):
            for col in ["B", "C"]:
                sla_data_cell_cord = f"{col}{row}"
                total_weight_cell_cord = f"{col}{total_weight_row_num}"
                # Every second row, chance the color.
                if row % 2 == 0:
                    self.fill_color(hex_color="e8f0fe", cell_coordinate=sla_data_cell_cord)

                self.create_full_borders(cell_coordinate=sla_data_cell_cord, border_type="thin")
                self.create_full_borders(cell_coordinate=total_weight_cell_cord, border_type="thin")
                self.change_font(cell_coordinate=total_weight_cell_cord, bold=True)
                self.change_font(cell_coordinate=sla_data_cell_cord, bold=False, hex_color="000000")
                self.fill_color(cell_coordinate=total_weight_cell_cord, hex_color='4285f4')

    def _common_header_design(self, data_dict: dict, size: int = 11, fill_hex_color: str = '4285f4',
                              font_hex_color: str = 'ffffff', bold: bool = True, italics: bool = False) -> None:
        """
        Change font, fill color and borders around common headers in the Excel File.
        :param data_dict: Dictionary of Data to use.
        :param size: Size you want the font. (Default: 11)
        :param font_hex_color: Color you want the font (Default: 'ffffff')
        :param fill_hex_color: Color you want the fill (Default: '4285f4')
        :param bold: If you want the font bold (Default: True)
        :param italics: If you want the font italic (Default: False)
        """
        header_coordinates = self.get_cell_coordinate([key for key in data_dict])
        for cell_cord in header_coordinates.keys():
            self.fill_color(cell_coordinate=cell_cord, hex_color=fill_hex_color)
            self.change_font(cell_coordinate=cell_cord, bold=bold, italics=italics, size=size, hex_color=font_hex_color)
            self.create_full_borders(cell_coordinate=cell_cord)

    def _sla_table_design(self) -> None:
        """
        Responsible for designing the SLA Table.
        """
        self._create_sla_headers()
        total_weight_row_num = self._add_sla_total_weight_header()
        self._sla_table_color(total_weight_row_num)

    def _other_design(self) -> None:
        """
        Designs that aren't in either SLA Table or Bot Table
        """
        self._add_color_reference_header()
        self._add_days_top_pri_header()
        self._add_date_header()
        self._add_logo(cell_coordinate="B2")

    def _bot_table_design(self) -> None:
        """
        Responsible for designing the Bot Table
        """
        self._common_header_design(data_dict=self.bot_df)

    def _create_bot_sla_report(self) -> None:
        """
        Creates the SLA/Bot Excel Report.
        """
        self._create_temp_file()

        # Convert sla_data to a Dataframe.
        sla_df = pd.DataFrame(self.sla_data, index=[0]).T

        # Insert Data
        self._insert_data_to_excel(dataframe=sla_df, start_row=9, start_col=2, header=False)
        self._insert_data_to_excel(dataframe=self.bot_df, start_row=8, start_col=5, index=False)

        # Design the Data
        self._open_temp_file()
        self._sla_table_design()
        self._bot_table_design()
        self._other_design()
        self.all_cell_styles()
        self.set_column_widths(column_widths=self._get_sla_bot_custom_widths())
        self.hide_gridlines()

        # Move/Save Excel File for User to view.
        self._create_excel_file(folder_name="SLA-Bot Report",
                                file_name=f"SLA-Bot Report on {ReportDesign.get_date_time()}.xlsx",
                                sheet_name="Bot Report")

    def _create_title_header(self, title_text: str) -> None:
        """
        Create the title header in the Home Delivery Report.
        :param title_text: Name of the title you want to display.
        """
        self.sheet.merge_cells("B3:C5")
        self.sheet["B3"].value = title_text
        self.fill_color(cell_coordinate="B3", hex_color="4285f4")
        self.change_font(cell_coordinate="B3", bold=True, italics=True, size=20)

        for row in self.sheet["B3:C5"]:
            for cell in row:
                self.create_full_borders(cell_coordinate=cell.coordinate)
                cell.alignment = Alignment(vertical="center", horizontal="center")

    def _shipped_awb_design(self) -> None:
        """
        Designs the Shipped AWB Sheet for Home Delivery Report.
        """
        self._common_header_design(self.shipped_awb_df, size=15)
        self.set_column_widths(column_widths=self._get_shipped_awb_custom_widths())
        self._create_title_header(title_text="Home Delivery Sent")
        self._add_logo(cell_coordinate='D1')
        self.hide_gridlines()

    def _non_shipped_awb_design(self) -> None:
        """
        Designs the Non-Shipped AWB Sheet for Home Delivery Report.
        """
        self._common_header_design(self.non_shipped_awb_df, size=15)
        self.set_column_widths(column_widths=self._get_non_shipped_awb_custom_widths())
        self._create_title_header(title_text="Home Delivery NOT Sent")
        self._add_logo(cell_coordinate='D1')
        self.hide_gridlines()

    def _create_home_delivery_report(self) -> None:
        """
        Creates the Home Delivery Excel Report.
        """
        self._create_temp_file()

        # Insert Data
        self._insert_data_to_excel(dataframe=self.shipped_awb_df, start_row=8, start_col=2, index=False)
        self._insert_data_to_excel(dataframe=self.non_shipped_awb_df, start_row=8, start_col=2, index=False,
                                  sheet_name="Non-Shipped AWB(s)")

        # Design the Data
        self._open_temp_file()

        # Change Sheets to Non-Shipped AWB(s)
        self.sheet = self.workbook["Non-Shipped AWB(s)"]
        self.all_cell_styles()
        self._non_shipped_awb_design()

        # Change Sheets to Non-Shipped AWB(s)
        self.sheet = self.workbook["Sheet"]
        self.all_cell_styles()
        self._shipped_awb_design()

        # Move/Save Excel File for User to view.
        self._create_excel_file(folder_name="Home Delivery Report",
                                file_name=f"Home Delivery Report on {ReportDesign.get_date_time()}.xlsx",
                                sheet_name="Shipped AWB(s)")

    def _open_temp_file(self) -> None:
        """
        Open the temporary file.
        """
        self.workbook = load_workbook(self.temp_file)
        self.sheet = self.workbook.active

    def _save_temp_file(self) -> None:
        """
        Save the temporary file
        """
        self.workbook.save(self.temp_file)

    def _create_temp_file(self) -> None:
        """
        Create a temporary Excel File
        """
        with tempfile.NamedTemporaryFile(suffix=".xlsx", delete=False) as temp:
            self.temp_file = temp.name
            self.workbook = Workbook()
            self.workbook.save(self.temp_file)

    @staticmethod
    def create_folder(folder_name) -> str:
        """
        Create a folder in the script directory.
        :param folder_name: Name of the folder.
        :return: Returns the path of the folder.
        """
        current_director = os.getcwd()
        folder_path = os.path.join(current_director, folder_name)

        if not os.path.isdir(folder_name):
            os.mkdir(folder_path)

        return folder_path

    @staticmethod
    def invalid_name(file_name: str = None, folder_name: str = None, sheet_name: str = None) -> None:
        """
        Checks if a file or folder name contain any illegal characters.
        :param file_name: Name of the file.
        :param folder_name: Name of the folder
        :param sheet_name: Name of the sheet
        :raise KeyError: Will raise error if file name, sheet name or folder name contain invalid characters. Also, will
            raise an error if file name doesn't contain .xlsx.
        """
        invalid_characters = ["\\", "/", ":", "*", "?", "<", ">", "|"]

        if file_name is not None:
            if any(char in invalid_characters for char in file_name):
                raise ValueError(
                    f'Invalid file name: {file_name}. The file name cannot contain {", ".join(invalid_characters)}')
            if ".xlsx" not in file_name:
                raise ValueError(
                    f'Invalid file name: {file_name}. The file name must contain the .xlsx file extension.')

        if folder_name is not None:
            if any(char in invalid_characters for char in folder_name):
                raise ValueError(
                    f'Invalid file name: {file_name}. The file name cannot contain {", ".join(invalid_characters)}')

        if sheet_name is not None:
            if any(char in invalid_characters[:4] for char in sheet_name):
                raise ValueError(f'Invalid sheet name: {sheet_name}. The sheet name cannot contain '
                                 f'{", ".join(invalid_characters[:3])}')

    def _create_excel_file(self, folder_name: str, file_name: str, sheet_name: str = None) -> None:
        """
        Creates a folder and moves the temporary Excel file into the folder.

        This will move the temporary file into the specific folder for the user to view/open.
        :param folder_name: Name of the folder to create. If folder is created, it will not create another one.
        :param file_name: Name of the Excel File.
        :param sheet_name: Name of the sheet. (Default: None)
        :raise KeyError: Will raise error if file name, sheet name or folder name contain invalid characters. Also, will
            raise an error if file name doesn't contain .xlsx.
        """
        if sheet_name is not None:
            self.sheet.title = sheet_name

        self._save_temp_file()

        ReportDesign.invalid_name(folder_name=folder_name, file_name=file_name, sheet_name=sheet_name)
        folder_path = ReportDesign.create_folder(folder_name)
        shutil.move(self.temp_file, f"{folder_path}/{file_name}")
